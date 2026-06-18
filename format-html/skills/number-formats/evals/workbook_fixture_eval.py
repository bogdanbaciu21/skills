#!/usr/bin/env python3
"""Workbook fixture eval for the number-formats skill.

Creates a small Excel workbook, runs apply-number-formats.py against it, then
reloads the workbook to verify formats and styles against formats.json.
"""

import json
import subprocess
import sys
import tempfile
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("workbook_fixture_eval.py needs openpyxl:  pip install openpyxl")


SKILL_DIR = Path(__file__).resolve().parents[1]
SCRIPT = SKILL_DIR / "apply-number-formats.py"
FORMATS = SKILL_DIR / "formats.json"


def font_rgb(cell):
    color = cell.font.color
    if color is None or color.type != "rgb":
        return None
    return color.rgb


def fill_rgb(cell):
    fill = cell.fill
    if fill is None or fill.fill_type != "solid" or fill.fgColor.type != "rgb":
        return None
    return fill.fgColor.rgb


def run_formatter(workbook, *args):
    cmd = [sys.executable, str(SCRIPT), str(workbook), *args, "--in-place"]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode:
        print(result.stdout, end="")
        print(result.stderr, end="", file=sys.stderr)
        sys.exit(result.returncode)


def run_help_without_site_packages():
    result = subprocess.run([sys.executable, "-S", str(SCRIPT), "--help"], capture_output=True, text=True)
    if result.returncode or "--auto-color" not in result.stdout:
        print(result.stdout, end="")
        print(result.stderr, end="", file=sys.stderr)
        sys.exit("formatter --help should not require openpyxl/site packages")
    if "--profile" not in result.stdout or "--used-range-format" not in result.stdout:
        sys.exit("formatter --help should expose whole-sheet helpers")


def build_fixture(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model"
    inputs = wb.create_sheet("Inputs")

    ws["B2"], ws["C2"], ws["D2"] = 1234.5, -1234.5, 0
    ws["B3"], ws["C3"], ws["D3"] = 0.125, -0.125, 0
    ws["B4"], ws["C4"], ws["D4"] = 1.5, -1.5, 0
    ws["B5"], ws["C5"], ws["D5"] = 1, 0, 2
    ws["B7"] = "=Inputs!A1"
    ws["B8"] = "=B2+B3"
    ws["B9"] = 500
    ws["B10"] = 1000
    ws["B11"] = 0.25
    inputs["A1"] = 42

    wb.save(path)


def expect(errors, label, actual, expected):
    if actual != expected:
        errors.append(f"{label}: expected {expected!r}, got {actual!r}")


def main():
    spec = json.loads(FORMATS.read_text(encoding="utf-8"))
    number_formats = spec["excel_number_formats"]
    provenance = spec["provenance_colors"]
    rows = spec["row_styles"]

    run_help_without_site_packages()

    with tempfile.TemporaryDirectory() as tmp:
        workbook = Path(tmp) / "number-format-fixture.xlsx"
        build_fixture(workbook)

        used_range_workbook = Path(tmp) / "used-range-fixture.xlsx"
        build_fixture(used_range_workbook)
        run_formatter(used_range_workbook, "--sheet", "Model", "--used-range-format", "number")
        used_range = openpyxl.load_workbook(used_range_workbook, data_only=False)["Model"]
        used_range_errors = []
        for address in ("B2", "C2", "D2", "B7", "B8", "B9"):
            expect(used_range_errors, f"{address} used-range format", used_range[address].number_format, number_formats["number"])
        if used_range_errors:
            print("number-formats used-range CLI eval failed:", file=sys.stderr)
            for error in used_range_errors:
                print(f"- {error}", file=sys.stderr)
            sys.exit(1)

        run_formatter(workbook, "--range", "Model!B2:D2", "--format", "number")
        run_formatter(workbook, "--range", "Model!B3:D3", "--format", "percent")
        run_formatter(workbook, "--range", "Model!B4:D4", "--format", "multiple")
        run_formatter(workbook, "--range", "Model!B5:D5", "--format", "yes-no")
        run_formatter(workbook, "--sheet", "Model", "--auto-color")
        run_formatter(workbook, "--range", "Model!B9", "--style", "input")
        run_formatter(workbook, "--range", "Model!B10", "--style", "total")
        run_formatter(workbook, "--range", "Model!B11", "--style", "margin")

        formatted = openpyxl.load_workbook(workbook, data_only=False)
        ws = formatted["Model"]

        errors = []
        for address in ("B2", "C2", "D2"):
            expect(errors, f"{address} number format", ws[address].number_format, number_formats["number"])
        for address in ("B3", "C3", "D3"):
            expect(errors, f"{address} percent format", ws[address].number_format, number_formats["percent"])
        for address in ("B4", "C4", "D4"):
            expect(errors, f"{address} multiple format", ws[address].number_format, number_formats["multiple"])
        for address in ("B5", "C5", "D5"):
            expect(errors, f"{address} toggle format", ws[address].number_format, number_formats["toggle_yes_no"])

        expect(errors, "hardcoded auto-color", font_rgb(ws["B2"]), provenance["input"]["font_argb"])
        expect(errors, "cross-sheet link color", font_rgb(ws["B7"]), provenance["link"]["font_argb"])
        expect(errors, "input style font", font_rgb(ws["B9"]), provenance["input"]["font_argb"])
        expect(errors, "input style fill", fill_rgb(ws["B9"]), provenance["input"]["fill_argb"])
        expect(errors, "total style bold", ws["B10"].font.bold, True)
        expect(errors, "margin style italic", ws["B11"].font.italic, True)
        expect(errors, "margin style color", font_rgb(ws["B11"]), rows["margin"]["font_argb"])

        run_formatter(workbook, "--sheet", "Model", "--profile", "operating-model")
        profiled = openpyxl.load_workbook(workbook, data_only=False)["Model"]
        expect(errors, "profile keeps numeric format", profiled["B2"].number_format, number_formats["number"])

        if errors:
            print("number-formats workbook eval failed:", file=sys.stderr)
            for error in errors:
                print(f"- {error}", file=sys.stderr)
            sys.exit(1)

    print("number-formats workbook eval passed")


if __name__ == "__main__":
    main()
