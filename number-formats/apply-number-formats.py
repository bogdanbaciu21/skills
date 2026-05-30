#!/usr/bin/env python3
"""apply-number-formats.py — apply the Macabacus number-format standard to a workbook.

Reads the byte-exact format codes + provenance colors from formats.json (generated
straight from the canonical source workbook) and applies them, so every model
formats numbers the same way: en-dash zeros, parenthesized negatives with aligned
decimals, $ only on the top row, blue inputs, green cross-sheet links, grey-italic
margins, bold totals.

Examples:
  # Number-format a range (number / percent / multiple / toggle):
  python3 apply-number-formats.py model.xlsx --range "Operating Model!F15:J40" --format number
  python3 apply-number-formats.py model.xlsx --range "Sheet1!F21:J21"          --format percent
  python3 apply-number-formats.py model.xlsx --range "Sheet1!F17:J17"          --format multiple
  python3 apply-number-formats.py model.xlsx --range "Sheet1!C2"               --format yes-no

  # Provenance / row styles:
  python3 apply-number-formats.py model.xlsx --range "Sheet1!C39"    --style input   # blue on yellow
  python3 apply-number-formats.py model.xlsx --range "Sheet1!F18:J18" --style total   # bold
  python3 apply-number-formats.py model.xlsx --range "Sheet1!F21:J21" --style margin  # grey italic

  # Auto-color by provenance across a sheet (hardcoded number -> blue input;
  # cross-sheet formula -> green link; other formulas left as default ink):
  python3 apply-number-formats.py model.xlsx --sheet "Operating Model" --auto-color

By default writes <file>-formatted.xlsx; pass --in-place to overwrite.
"""
import argparse, json, os, re, sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    sys.exit("apply-number-formats.py needs openpyxl:  pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
SPEC = json.load(open(os.path.join(HERE, "formats.json")))
NF, PROV, ROW = SPEC["excel_number_formats"], SPEC["provenance_colors"], SPEC["row_styles"]

FORMAT_ALIAS = {
    "number": "number", "num": "number", "n": "number",
    "percent": "percent", "pct": "percent", "%": "percent",
    "multiple": "multiple", "mult": "multiple", "x": "multiple",
    "yes-no": "toggle_yes_no", "y-n": "toggle_y_n", "on-off": "toggle_on_off",
}


def parse_ref(ref, default_sheet):
    if "!" in ref:
        sheet, rng = ref.split("!", 1)
        return sheet.strip("'"), rng
    return default_sheet, ref


def cells_in(ws, rng):
    sel = ws[rng]
    if hasattr(sel, "value"):          # single cell
        yield sel
        return
    for row in sel:                    # range of rows
        for c in row:
            yield c


def _font(c, **kw):
    f = c.font
    base = dict(name=f.name, size=f.size, bold=f.bold, italic=f.italic, color=f.color)
    base.update(kw)
    return Font(**base)


def apply_format(ws, rng, key):
    n = 0
    for c in cells_in(ws, rng):
        c.number_format = NF[key]
        n += 1
    return n


def apply_style(ws, rng, style):
    n = 0
    for c in cells_in(ws, rng):
        if style in ("input", "hardcode", "link", "error"):
            c.font = _font(c, color=PROV[style]["font_argb"])
            if style == "input" and PROV["input"].get("fill_argb"):
                c.fill = PatternFill("solid", fgColor=PROV["input"]["fill_argb"])
        elif style == "total":
            c.font = _font(c, bold=True)
        elif style == "margin":
            c.font = _font(c, italic=True, color=ROW["margin"]["font_argb"])
        else:
            sys.exit(f"unknown style '{style}'")
        n += 1
    return n


def auto_color(ws):
    inputs = links = 0
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None:
                continue
            if isinstance(v, str) and v.startswith("="):
                if re.search(r"[A-Za-z0-9_']+!", v):          # references another sheet
                    c.font = _font(c, color=PROV["link"]["font_argb"]); links += 1
                # plain formula -> leave default ink
            elif isinstance(v, (int, float)):
                c.font = _font(c, color=PROV["input"]["font_argb"]); inputs += 1
    return inputs, links


def main():
    ap = argparse.ArgumentParser(description="Apply the Macabacus number-format standard.")
    ap.add_argument("file")
    ap.add_argument("--range", help="A1 range, optionally qualified: 'Sheet!A1:B2'")
    ap.add_argument("--format", dest="fmt", help="number | percent | multiple | yes-no | y-n | on-off")
    ap.add_argument("--style", help="input | hardcode | link | error | total | margin")
    ap.add_argument("--sheet", help="default sheet name (else the active sheet)")
    ap.add_argument("--auto-color", action="store_true", help="color cells by provenance")
    ap.add_argument("--in-place", action="store_true", help="overwrite the file instead of *-formatted.xlsx")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.file)
    default_sheet = args.sheet or wb.active.title
    did = []

    if args.auto_color:
        ws = wb[default_sheet]
        i, l = auto_color(ws)
        did.append(f"auto-color '{ws.title}': {i} inputs -> blue, {l} cross-sheet links -> green")
    if args.range and args.fmt:
        sheet, rng = parse_ref(args.range, default_sheet)
        key = FORMAT_ALIAS.get(args.fmt.lower())
        if not key:
            sys.exit(f"unknown --format '{args.fmt}' (number|percent|multiple|yes-no|y-n|on-off)")
        did.append(f"format {key}: {sheet}!{rng} ({apply_format(wb[sheet], rng, key)} cells)")
    if args.range and args.style:
        sheet, rng = parse_ref(args.range, default_sheet)
        did.append(f"style {args.style}: {sheet}!{rng} ({apply_style(wb[sheet], rng, args.style)} cells)")

    if not did:
        sys.exit("nothing to do — pass --range with --format/--style, or --auto-color")

    out = args.file if args.in_place else os.path.splitext(args.file)[0] + "-formatted.xlsx"
    wb.save(out)
    print("\n".join(did))
    print("saved:", out)


if __name__ == "__main__":
    main()
